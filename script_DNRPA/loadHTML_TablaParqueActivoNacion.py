import datetime
import mysql.connector
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl
import pandas as pd
from datetime import datetime
from selenium.webdriver.support.ui import Select
import os


host = 'localhost'
user = 'root'
password = 'Estadistica123'
database = 'prueba1'

class loadHTML_TablaParqueActivoNacion:
    def loadInDataBase(self, host, user, password, database):
        # Se toma el tiempo de comienzo
        start_time = time.time()

        # Establecer la conexión a la base de datos
        conn = mysql.connector.connect(
            host=host, user=user, password=password, database=database
        )
        try:
            # Obtener la ruta del directorio actual (donde se encuentra el script)
            directorio_actual = os.path.dirname(os.path.abspath(__file__))
            ruta_carpeta_files = os.path.join(directorio_actual, 'files')
            ruta_archivo_excel = os.path.join(ruta_carpeta_files, 'registroParqueActivoNacion.xlsx')


            driver = webdriver.Chrome()
            driver.get('https://www.dnrpa.gov.ar/portal_dnrpa/boletines_estadisticos2.php')

            # Obtener la ventana actual
            ventana_actual = driver.current_window_handle
            
            elemento = driver.find_element(By.XPATH, '/html/body/div/section/article/div/div[2]/div/div/div/div/form/table/tbody/tr[2]/td[2]/select')

            # Localizar el elemento select por su nombre
            select_element = elemento.find_element(By.XPATH, '/html/body/div/section/article/div/div[2]/div/div/div/div/form/table/tbody/tr[2]/td[2]/select')

            # Crear un objeto Select
            select = Select(select_element)
            
            # O seleccionar por texto visible
            valor_deseado = '2023'
            select.select_by_visible_text('Año 2023')
            
            # Esperar un momento para que se abra la nueva pestaña
            driver.implicitly_wait(5)
            # Cambiar al contexto de la nueva pestaña
            for ventana in driver.window_handles:
                if ventana != ventana_actual:
                    driver.switch_to.window(ventana)
            
            time.sleep(5)
           #↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓ ENCONTRAR Y TOMAR LOS DATOS ↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓
            # Encontrar el elemento <div> con la clase 'grid'
            elemento_div = driver.find_element(By.CLASS_NAME, 'grid')

            # Encontrar la tabla dentro del elemento <div>
            elemento_tabla = elemento_div.find_element(By.TAG_NAME, 'table')

            # Obtener todas las filas de la tabla
            filas = elemento_tabla.find_elements(By.TAG_NAME, 'tr')

            # Lista para almacenar los datos de la tabla
            tabla_datos = []
            
            for fila in filas:
                # Obtener las celdas de cada fila, excluyendo la última columna y la última celda de encabezado
                celdas = fila.find_elements(By.TAG_NAME, 'th') + fila.find_elements(By.TAG_NAME, 'td')

                # Lista para almacenar los valores de cada fila
                fila_datos = []

                for celda in celdas:
                    valor = celda.text
                    if isinstance(valor, str):
                        # Verificar si el valor comienza con un número
                        if valor.strip() and valor[0].isdigit():
                            try:
                                # Reemplazar el punto decimal por una coma (si es necesario)
                                valor = valor.replace('.', '')
                                # Intentar convertir el valor a float
                                valor = float(valor)
                                print("valor1: ", valor)
                            except ValueError:
                                pass  # Mantener el valor original si no se puede convertir a float
                    fila_datos.append(valor)
                print("aca: ", fila_datos)


                tabla_datos.append(fila_datos) 
            
            datos_sin_segunda_fila = tabla_datos[0:1] + tabla_datos[2:]
            # Transponer los datos utilizando pandas
            df = pd.DataFrame(datos_sin_segunda_fila)
            df_transpuesta = df.transpose()
            
            #Conversion de MESES a formato Y-M-D , tipo de dato: datetime
            print(df_transpuesta[0][1:])
            meses = df_transpuesta[0][1:]

            #Donde almacenamos las nuevas fechas
            nuevas_fechas = list()

            for i in range(1, len(meses)+ 1):

                if i < 10:
                    fecha_str =  '01-0'+str(i)+"-"+ str(valor_deseado)
                else:
                    fecha_str = '01-'+str(i)+"-"+ str(valor_deseado)

                fecha_str = datetime.strptime(fecha_str,'%d-%m-%Y').date()
                nuevas_fechas.append(fecha_str)

            #Reasignacion de fechas
            df_transpuesta[0][1:] = nuevas_fechas
            
            # Convertir los valores de la transposición a una lista
            valores_transpuestos = df_transpuesta.values.tolist()
            
        #↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓ EXCEL ↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓
            # Cargar el archivo Excel existente
            libro_excel = openpyxl.load_workbook(ruta_archivo_excel)

           # Seleccionar la hoja activa del libro
            hoja_activa = libro_excel.active

            # Obtener la cantidad total de filas en el archivo
            total_filas = hoja_activa.max_row

            # Eliminar todas las filas excepto la primera (encabezados)
            hoja_activa.delete_rows(1, total_filas)
            
            # Recorrer los datos transpuestos y escribirlos en el archivo de Excel
            for fila_datos in valores_transpuestos:
                hoja_activa.append(fila_datos)
                
            df_transpuesta.drop

            # Guardar el archivo Excel actualizado sobreescribiendo los datos existentes
            libro_excel.save(ruta_archivo_excel)
            
            #↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓ CARGA MYSQL ↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓
            # Crear el cursor para ejecutar consultas
            cursor = conn.cursor()
            # Cargar el archivo Excel
           
            # Cargar el archivo Excel
            workbook = openpyxl.load_workbook(ruta_archivo_excel)

            column_mapping = {
                'Provincia / Mes': 'Fecha',
                'BUENOS AIRES': 'Buenos_Aires',
                'CABA': 'C_Autonoma_De_BSAS',
                'CATAMARCA': 'Catamarca',
                'CORDOBA': 'Cordoba',
                'CORRIENTES': 'Corrientes',
                'CHACO': 'Chaco',
                'CHUBUT': 'Chubut',
                'ENTRE RIOS': 'Entre_Rios',
                'FORMOSA': 'Formosa',
                'JUJUY': 'Jujuy',
                'LA PAMPA': 'La_Pampa',
                'LA RIOJA': 'La_Rioja',
                'MENDOZA': 'Mendoza',
                'MISIONES': 'Misiones',
                'NEUQUEN': 'Neuquen',
                'RIO NEGRO': 'Rio_Negro',
                'SALTA': 'Salta',
                'SAN JUAN': 'San_Juan',
                'SAN LUIS': 'San_Luis',
                'SANTA CRUZ': 'Santa_Cruz',
                'SANTA FE': 'Santa_Fe',
                'S. DEL ESTERO': 'Sgo_Del_Estero',
                'TUCUMAN': 'Tucuman',
                'T. DEL FUEGO': 'Tierra_Del_Fuego',
                'TOTAL': 'Total_Nacion'
            }

            # Obtener la hoja de trabajo específica
            sheet = workbook["Hoja1"]

            # Obtener las fechas existentes en la tabla de MySQL
            select_dates_query = "SELECT Fecha FROM dnrpa_parque_activo_nacion"
            cursor.execute(select_dates_query)
            existing_dates = [row[0].strftime('%Y-%m-%d') for row in cursor.fetchall()]
            
            # Recorrer las filas del archivo de Excel a partir de la segunda fila
            for row in sheet.iter_rows(min_row=2, values_only=True):
                fecha = row[0]
                fecha_str = fecha.strftime('%Y-%m-%d')
                
                if fecha_str in existing_dates:
                    # Realizar una actualización (UPDATE)
                    update_values = []

                    # Recorrer las columnas y los valores de la fila actual
                    for col_idx, value in enumerate(row):
                        # Obtener el nombre de la columna correspondiente en el archivo de Excel
                        column_name_excel = sheet.cell(row=1, column=col_idx + 1).value

                        # Verificar si la columna está mapeada en el diccionario de mapeo
                        if column_name_excel in column_mapping:
                            # Obtener el nombre de la columna correspondiente en la base de datos MySQL
                            column_name_mysql = column_mapping[column_name_excel]

                            # Agregar el nombre de la columna y el valor a la lista de valores para la actualización
                            update_values.append((column_name_mysql, value))

                    # Crear la sentencia SQL para la actualización
                    update_query = "UPDATE dnrpa_parque_activo_nacion SET " + ", ".join([f"{col[0]} = %s" for col in update_values]) + " WHERE Fecha = %s"
                    # Obtener los valores de la columna en el orden correcto para la actualización
                    update_values = [col[1] for col in update_values]

                    # Agregar la fecha al final de los valores de actualización
                    update_values.append(fecha)

                    # Ejecutar la sentencia SQL
                    cursor.execute(update_query, update_values)
                else:
                    # Realizar una inserción (INSERT)
                    insert_values = []

                    # Recorrer las columnas y los valores de la fila actual
                    for col_idx, value in enumerate(row):
                        # Obtener el nombre de la columna correspondiente en el archivo de Excel
                        column_name_excel = sheet.cell(row=1, column=col_idx + 1).value

                        # Verificar si la columna está mapeada en el diccionario de mapeo
                        if column_name_excel in column_mapping:
                            # Obtener el nombre de la columna correspondiente en la base de datos MySQL
                            column_name_mysql = column_mapping[column_name_excel]

                            # Agregar el nombre de la columna y el valor a la lista de valores para la inserción
                            insert_values.append((column_name_mysql, value))

                    # Crear la sentencia SQL para la inserción
                    columns = ", ".join([col[0] for col in insert_values])
                    placeholders = ", ".join(["%s" for _ in insert_values])
                    insert_query = f"INSERT INTO dnrpa_parque_activo_nacion ({columns}) VALUES ({placeholders})"

                    # Obtener los valores de la columna en el orden correcto para la inserción
                    insert_values = [col[1] for col in insert_values]

                    # Ejecutar la sentencia SQL
                    cursor.execute(insert_query, insert_values)

            # Confirmar los cambios en la base de datos
            conn.commit()
            
            # Se toma el tiempo de finalización y se calcula
            end_time = time.time()
            duration = end_time - start_time
            print("-----------------------------------------------")
            print("Se guardaron los datos de Registro automotor")
            print("Tiempo de ejecución:", duration)

            # Cerrar la conexión a la base de datos
            conn.close()

        except Exception as e:
            # Manejar cualquier excepción ocurrida durante la carga de datos
            print(f"Registro automotor: Ocurrió un error durante la carga de datos: {str(e)}")
            conn.close()  # Cerrar la conexión en caso de error

loadHTML_TablaParqueActivoNacion().loadInDataBase(host, user, password, database)